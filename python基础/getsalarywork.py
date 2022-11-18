# -*- coding: utf-8 -*-
"""
====================================
@File Name ：getsalarywork.py
@Time ： 2022/11/17 14:26
@Program IDE ：PyCharm
@Create by Author ： stone
@Describe：通过读取xlsx的获取工资信息
====================================
"""
import os
import openpyxl
from datetime import timedelta
import pandas as pd
from utils.logger import Logger

file_dir = os.path.dirname(os.path.abspath(__file__)) + "\\file\\"
log_dir = os.path.dirname(os.path.abspath(__file__)) + "\\log\\"
logger = Logger(log_dir, os.path.basename(__file__))


def read_xlsx(xlsx_name):
    dir_path = os.path.join(file_dir, xlsx_name)
    # 判断文件路径是否存在
    if not os.path.exists(dir_path):
        logger.debug("%s路径不存在" % dir_path)
        return
    # wb = openpyxl.load_workbook(dir_path)
    excel_reader = pd.ExcelFile(dir_path)
    # logger.debug("{}有sheet页{}".format(xlsx_name, wb.sheetnames))
    # sheets = wb.sheetnames
    # for sheet in sheets:
    #     logger.debug("当前遍历的sheet{}".format(sheet))
    #     ws = wb[sheet]
    #     df = pd.DataFrame(ws.values)
    #     print(df)
    #     #当前页面的标头
    #     print(df.columns.tolist())
    sheet_names = excel_reader.sheet_names
    for sheet_name in sheet_names:
        logger.debug("当前遍历的sheet{}".format(sheet_name))
        df = excel_reader.parse(sheet_name)
        # 当前页面的标头
        list_name = df.columns.tolist()
        for i, row in df.iterrows():
            # 封装当前用的户信息
            user_name = row["用户名"]
            salar = row["工资"]
            email = row['邮箱']
            # 针对签到时间的处理
            # print(row["签到时间"])
            work = work_time_anlays(row["签到时间"])
            body = "{}发工资{}元,工作时间{}分钟".format(user_name, salar,work/60)
            logger.info("发送邮件给{}内容为{}".format(email, body))
            # 获取当前用户的邮箱


def work_time_anlays(worktime):
    if worktime is None:
        return 0
    work_time_list = str(worktime).split('\n')
    if len(work_time_list) <= 1:
        return 0
    start_time = work_time_list[0].strip()
    end_time = work_time_list[-1].strip()
    stime = timedelta(hours=int(start_time.split(":")[0]), minutes=int(start_time.split(":")[1]))
    etime = timedelta(hours=int(end_time.split(":")[0]), minutes=int(end_time.split(":")[1]))
    w_time = etime - stime
    return w_time.seconds


if __name__ == '__main__':
    salary_name = "salary.xlsx"
    read_xlsx(salary_name)
